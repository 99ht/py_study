from openpyxl import load_workbook
import os
# 加载工作簿
workbook = load_workbook(filename='工作簿1.xlsx')

# 选择工作表
sheet = workbook['Sheet1']

# # 读取单元格内容
# cell_value = sheet['A1'].value
# print(f"A1单元格的值是：{cell_value}")
#
# # 读取整列数据
# row_values = [cell.value for cell in sheet['A']]
# print(f"A列的值是：{row_values}")
#
# # 读取整行数据
column_values = [cell.value for cell in sheet[1]]
print(f"第1行的值是：{column_values}")
print(sheet[1][2])


# 读取整个工作表数据
# for row in sheet.iter_rows(max_row=7, max_col=5, values_only=True):
#
#     for cell in row:
#         print(cell)
#         print(type(cell))
curr_dir = os.getcwd()
tar_file_path = curr_dir + r"/mass.h"
if os.path.exists(tar_file_path):
    os.remove(tar_file_path)


def get_cell_value(sheet_num, row_num, column_num):
    sheet_value = next(sheet_num.iter_rows(min_row=row_num, max_row=row_num, min_col=column_num, max_col=column_num, values_only=True))
    if sheet_value[0] is not None:
        return sheet_value[0]


with open(tar_file_path, "w") as tar_file:
    for row in range(3, sheet.max_row + 1):
        # for column in range(1, sheet.max_column + 1):
        # name = get_cell_value(sheet, row, 6)
        name = sheet.cell(row, 6).value
        if name is None:
            continue
        value = sheet.cell(row, 5).value
        note = sheet.cell(row, 3).value

        # 获取特定单元格的值（例如 A1）
        # cell_value = sheet['A1'].value

        # value = get_cell_value(sheet, row, 5)
        # note = get_cell_value(sheet, row, 3)
        tar_file.write("#define " + name + " " + value + " // " + note + "\r\n")
