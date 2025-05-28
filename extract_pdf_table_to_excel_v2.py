import pdfplumber
import openpyxl as op
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import os
from datetime import datetime
import re

# 设置全局单元格样式
alignment_style_auto_LF = Alignment(
    horizontal='left',  # 水平靠左
    vertical='center',  # 垂直居中
    wrap_text=True  # 自动换行（确保多行内容可见）
)

settings = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines",
    "snap_tolerance": 6,  # 增大捕捉容差
    "join_tolerance": 10,  # 放宽合并阈值
    "edge_min_length": 15,  # 忽略短线段
    "text_x_tolerance": 6,  # 文字水平偏移容差
    "text_y_tolerance": 6,
    "intersection_y_tolerance": 10
}

# 定义边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def is_row_empty(sheet, row_idx):
    #"""检查指定行是否为空（包括空字符串和None）"""
    for cell in sheet[row_idx]:
        if cell.value not in [None, "", " "]:  # 排除None、空字符串和空格
            return False
    return True

def auto_fit_columns(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            # 计算列最大内容宽度（考虑中文）
            for cell in column:
                try:
                    #cell_value = str(cell.value) if cell.value else ""
                    cell_value = re.split(r'\r?\n', cell.value, 1)[0] if cell.value else ""
                    # 中文字符按2单位计算，英文按1单位
                    length = sum(2 if ord(c) > 255 else 1 for c in cell_value)
                    max_length = max(max_length, length)
                except:
                    pass

            # 设置列宽（加2单位缓冲）
            adjusted_width = (max_length + 2)* 1.03  # 1.03是字体宽度系数
            ws.column_dimensions[column_letter].width = min(adjusted_width, 120)  # 限制最大宽度
            # 应用到所有单元格
            for row in ws.iter_rows():
                for cell in row:
                    #if cell.value: #检查单元格存储的​​值​​是否为空
                    #if cell is not None: #检查​​单元格对象​​是否存在
                    cell.alignment = alignment_style_auto_LF
                    
            for row_idx in range(1, ws.max_row + 1):
                if not is_row_empty(ws, row_idx):
                    row = ws[row_idx]  #获取当前行对象
                    for cell in row:
                        cell.border = thin_border  #给单元格加框线
    wb.save(file_path)


def extract_and_merge_pdf_tables(pdf_path):
    all_tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            #print(page_num)
            # 提取当前页的所有表格
            tables = page.extract_tables(settings)
            for idx, table in enumerate(tables):  #对tables 列表里的每个表格进行遍历
                ws2 = wb.create_sheet(title=f"Page{page_num + 1}_Table{idx + 1}")
                merged_table = []
                for row in table:  #对当前表格里的每一行进行遍历
                    #print(len(row)) #列数
                    merged_row = []
                    for cell in row:  #对当前行里的每个单元格进行遍历
                        if cell is not None:
                            merged_cell = cell
                            merged_row.append(merged_cell)
                    ws2.append(merged_row)
                    merged_table.append(merged_row)
                all_tables.append(merged_table)
    return all_tables


#读取当前目录所有pdf文件
def get_all_pdf_files():
    current_directory = os.getcwd()
    pdf_files = []
    #print(os.listdir(current_directory))
    for file in os.listdir(current_directory):
        if file.lower().endswith('.pdf'):
            pdf_file_path = os.path.join(current_directory, file)
            pdf_files.append(pdf_file_path)
    return pdf_files


pdf_files = get_all_pdf_files()
for pdf_file in pdf_files:
    wb = op.Workbook()
    ws = wb.active  # 创建子表
    #ws.title = "sheet"
    tables = extract_and_merge_pdf_tables(pdf_file)
    if tables:  #没有table就不保存excel文件了
        wb.remove(ws)
        output_excel_path = os.path.splitext(pdf_file)[0] + '.xlsx'  #需要生成的文件名
        curr_year = datetime.now().year
        wb.save(output_excel_path)  # 替换为你想要保存的 Excel 文件路径
        # 使用示例
        auto_fit_columns(output_excel_path)
